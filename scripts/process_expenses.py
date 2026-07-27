import json
import os
import re
import subprocess
import argparse
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output"
EXPENSE_SUFFIX = "_20260702130555.xlsx"


def find_expense_file():
    for root, _, files in os.walk("D:/"):
        for name in files:
            if name.endswith(EXPENSE_SUFFIX) and not name.startswith("~$"):
                return Path(root) / name
    raise FileNotFoundError(EXPENSE_SUFFIX)


def extract_budget_data():
    js = r"""
const fs=require('fs');
const file=fs.readdirSync('.').find(f=>f.endsWith('v2.html'));
const txt=fs.readFileSync(file,'utf8');
function extractVar(name){
  const start=txt.indexOf('var '+name+'=');
  if(start<0) throw new Error('missing '+name);
  let i=txt.indexOf('=',start)+1, depth=0, inStr=false, quote='', esc=false;
  for(let j=i;j<txt.length;j++){
    const ch=txt[j];
    if(inStr){ if(esc) esc=false; else if(ch==='\\') esc=true; else if(ch===quote) inStr=false; continue; }
    if(ch==='"'||ch==="'"){inStr=true; quote=ch; continue;}
    if(ch==='['||ch==='{') depth++;
    if(ch===']'||ch==='}') depth--;
    if(depth===0 && ch===';') return txt.slice(i,j);
  }
}
function I(id,g,n,c,u,q,o,a,e){return{id,group:g,name:n,calc:c,unitPrice:u,qty1:q,original:o,adjusted:a,executed:e};}
const initialCourses=eval(extractVar('initialCourses'));
console.log(JSON.stringify(initialCourses));
"""
    result = subprocess.run(
        ["node", "-e", js],
        cwd=ROOT,
        check=True,
        text=True,
        encoding="utf-8",
        capture_output=True,
    )
    return json.loads(result.stdout)


def load_base_courses(base_json):
    if not base_json:
        return extract_budget_data()
    path = Path(base_json)
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    courses = payload.get("data", {}).get("courses")
    if not isinstance(courses, list):
        raise ValueError(f"{path} 파일에 data.courses 구조가 없습니다.")
    return courses


def norm(text):
    return re.sub(r"\s+", "", str(text or "")).lower()


COURSE_PATTERNS = [
    (1, ["행정회계사무oa과정(1기)", "행정회계사무oa(1기)", "행정회계사무oa과정1기", "사무oa과정(1기)", "사무oa(1기)"]),
    (2, ["행정회계사무oa과정(2기)", "행정회계사무oa(2기)", "행정회계사무oa과정2기", "사무oa과정(2기)", "사무oa(2기)"]),
    (3, ["haccp전문인력양성과정", "haccp전문인력양성", "haccp"]),
    (4, ["erp·지게차물류관리실무자양성과정", "erp지게차물류관리실무자양성과정", "지게차물류관리", "erp·지게차"]),
    (5, ["승강기전문가양성과정", "승강기전문가양성"]),
    (6, ["중장년기회강사양성과정", "중장년기회강사양성"]),
    (7, ["초등피지컬코딩강사양성과정", "초등피지컬코딩강사양성", "피지컬코딩"]),
    (8, ["ai딥러닝전문강사양성과정", "ai딥러닝전문강사양성", "딥러닝전문강사"]),
    (9, ["늘봄학교창의융합교육강사양성과정", "늘봄학교창의융합교육강사양성", "창의융합교육강사"]),
    (10, ["ai시대의캐릭터크리에이터양성과정", "ai시대의캐릭터크리에이터양성", "캐릭터크리에이터"]),
]


def detect_course(content):
    n = norm(content)
    for course_id, patterns in COURSE_PATTERNS:
        if any(p in n for p in patterns):
            return course_id
    return 0


def detect_item(account, content, vendor, course):
    n = norm(content)
    v = norm(vendor)

    if course == 0:
        if "카카오톡" in n or "슈어엠" in n or "문자" in n or "sns" in n:
            return "SNS·문자 홍보"
        if "홍보" in n or "홍보물" in n:
            return "교육과정 홍보물 제작"
        if "직종설명회" in n:
            return "직종설명회 운영"
        if "간담회" in n:
            return "강사 간담회 운영"
        if "다과" in n or "소모품" in n:
            return "다과 및 소모품"
        if "특강" in n or "강사" in n:
            return "심화교육 특강 강사료"
        return None

    if "면접수당" in n or "심사수당" in n or "면접" in n:
        return "심사수당"
    if "홍보물" in n or "현수막" in n or "모집홍보" in n or "교육생모집" in n:
        return "모집홍보비"
    if "우편" in n or "발송" in n:
        return "우편발송료"
    if "훈련수당" in n or "훈련지원금" in n:
        return "훈련지원금" if course in {1, 2, 3, 4, 9} else "훈련수당"
    if "학습동아리" in n:
        return "학습동아리"
    if "취업" in n and "특강" in n:
        if course == 4:
            return "취업특강 및 안전교육"
        if course == 10:
            return "취업·저작권 특강"
        return "취업특강" if course in {7, 8, 9} else "취업대비특강"
    if "생성형ai" in n:
        return "생성형AI 특강"
    if "ai" in n and "특강" in n:
        return "AI 특강"
    if "사후관리" in n:
        return "사후관리특강"
    if "강사료" in n or "강사수당" in n or ("강사" in n and "수당" in n):
        if course == 4 and "erp" in n:
            return "ERP 강사수당"
        if course == 4:
            return "강사수당(지게차)"
        if course == 6:
            return "강사료"
        return "강사수당"
    if "보조강사" in n:
        return "보조강사"
    if "실습용역" in n:
        return "실습용역"
    if "실험보조" in n:
        return "실험보조"
    if "장비사용" in n or "교육시설임차" in n or "교육장임차" in n:
        return "교육장 임차료"
    if "버스임차" in n:
        return "버스 임차료"
    if "보험" in n:
        return "보험료"
    if "승강기길잡이교육비" in n or "실습교육" in n:
        return "실습 교육 비용"
    if "교육운영" in n:
        return "교육 운영비"
    if "교구" in n or "재료" in n:
        if course in {7, 8}:
            return "재료비"
        if course == 9:
            return "교재 및 재료"
    if "교재" in n or "책자" in n:
        if course == 6 and "mbti" in n:
            return "교재(MBTI 검사지)"
        return "교재비용" if course in {1, 2} else "교재비"
    if "소모품" in n:
        return "소모품비"
    if "구루미" in n:
        return "구루미 사용"
    if "프로필" in n:
        return "강사프로필"
    if "강사소개" in n or "소개책자" in n:
        return "강사소개책자"
    return None


def find_item(course, item_name, account):
    active = [item for item in course["items"] if not item.get("isDeleted")]
    for item in active:
        if item["name"] == item_name:
            return item
    candidates = [item for item in active if item["group"] == account]
    if len(candidates) == 1:
        return candidates[0]
    return None


def to_iso(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "")


def auto_width(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for cell in ws[letter]:
            width = max(width, min(60, len(str(cell.value or "")) + 2))
        ws.column_dimensions[letter].width = width


def write_sheet(ws, rows, headers):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    auto_width(ws)


def main():
    parser = argparse.ArgumentParser(description="지출내역을 예산프로그램 데이터 구조에 맞춰 정리합니다.")
    parser.add_argument(
        "--base-json",
        help="조정 후 예산프로그램에서 내보낸 JSON. 지정하면 조정예산/항목 구조를 유지하고 집행액만 다시 반영합니다.",
    )
    args = parser.parse_args()

    expense_path = find_expense_file()
    courses = load_base_courses(args.base_json)
    course_by_id = {course["id"]: course for course in courses}
    item_lookup = {}

    wb = openpyxl.load_workbook(expense_path, data_only=True)
    ws = wb.active

    raw_rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[9] is None:
            continue
        account = row[4]
        content = row[7]
        vendor = row[8]
        amount = int(round(float(row[9] or 0)))
        course_id = detect_course(content)
        item_name = detect_item(account, content, vendor, course_id)
        item = find_item(course_by_id[course_id], item_name, account) if item_name else None
        confidence = "높음" if course_id != 0 and item else "보통" if item else "확인필요"
        raw_rows.append(
            {
                "source_row": idx,
                "course_id": course_id,
                "course": course_by_id[course_id]["name"],
                "account": account,
                "item_id": item["id"] if item else "",
                "item": item["name"] if item else (item_name or "미분류_확인필요"),
                "date": to_iso(row[6] or row[5]),
                "content": content,
                "vendor": vendor,
                "amount": amount,
                "resolution_no": row[10],
                "evidence": row[11],
                "note": row[13],
                "confidence": confidence,
            }
        )

    grouped = defaultdict(int)
    for row in raw_rows:
        if row["item_id"]:
            grouped[(row["course_id"], row["item_id"])] += row["amount"]

    executions = []
    for i, row in enumerate(raw_rows, start=1):
        if not row["item_id"]:
            continue
        executions.append(
            {
                "id": 20260702000000 + i,
                "courseId": row["course_id"],
                "itemId": row["item_id"],
                "date": row["date"],
                "amount": row["amount"],
                "vendor": row["vendor"],
                "memo": f"{row['content']} / 결의 {row['resolution_no'] or ''}".strip(),
            }
        )

    for course in courses:
        for item in course["items"]:
            item["executed"] = grouped.get((course["id"], item["id"]), 0)

    OUTPUT_DIR.mkdir(exist_ok=True)
    json_path = OUTPUT_DIR / "예산프로그램_지출반영_20260702.json"
    payload = {
        "version": 2,
        "savedAt": datetime.now().isoformat(timespec="seconds"),
        "data": {
            "courses": courses,
            "executions": executions,
            "logs": [
                {
                    "id": "import-20260702",
                    "reason": "지출내역_20260702130555.xlsx 기준 집행액 반영",
                    "editedAt": datetime.now().isoformat(timespec="seconds"),
                    "editedBy": "Codex",
                    "after": {"rows": len(raw_rows), "matchedRows": len(executions)},
                }
            ],
        },
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    summary_rows = []
    total_original = total_adjusted = total_executed = 0
    for course in courses:
        for item in course["items"]:
            original = int(item.get("original") or 0)
            adjusted = int(item.get("adjusted") or 0)
            executed = int(item.get("executed") or 0)
            if original or adjusted or executed:
                total_original += original
                total_adjusted += adjusted
                total_executed += executed
                summary_rows.append(
                    [
                        course["name"],
                        course["manager"],
                        course["category"],
                        item["group"],
                        item["name"],
                        original,
                        adjusted,
                        executed,
                        adjusted - executed,
                        round(executed / adjusted * 100, 1) if adjusted else 0,
                    ]
                )

    account_rows = []
    by_account = defaultdict(lambda: [0, 0, 0])
    by_course_account = defaultdict(lambda: [0, 0, 0])
    for course in courses:
        for item in course["items"]:
            values = [int(item.get("original") or 0), int(item.get("adjusted") or 0), int(item.get("executed") or 0)]
            by_account[item["group"]][0] += values[0]
            by_account[item["group"]][1] += values[1]
            by_account[item["group"]][2] += values[2]
            by_course_account[(course["id"], item["group"])][0] += values[0]
            by_course_account[(course["id"], item["group"])][1] += values[1]
            by_course_account[(course["id"], item["group"])][2] += values[2]
    for account, values in sorted(by_account.items()):
        account_rows.append([account, *values, values[1] - values[2], round(values[2] / values[1] * 100, 1) if values[1] else 0])

    course_rows = []
    for course in courses:
        original = sum(int(item.get("original") or 0) for item in course["items"])
        adjusted = sum(int(item.get("adjusted") or 0) for item in course["items"])
        executed = sum(int(item.get("executed") or 0) for item in course["items"])
        course_rows.append([course["name"], course["manager"], course["category"], original, adjusted, executed, adjusted - executed, round(executed / adjusted * 100, 1) if adjusted else 0])

    account_order = ["사무관리비", "공공운영비", "행사운영비", "회의비", "교육훈련비", "행사실비보상금", "보상금"]
    course_account_rows = []
    for course in courses:
        for account in account_order:
            values = by_course_account.get((course["id"], account), [0, 0, 0])
            if not any(values):
                continue
            course_account_rows.append(
                [
                    course["name"],
                    course["manager"],
                    course["category"],
                    account,
                    values[0],
                    values[1],
                    values[2],
                    values[1] - values[2],
                    round(values[2] / values[1] * 100, 1) if values[1] else 0,
                ]
            )

    course_account_matrix = []
    for course in courses:
        row = [course["name"], course["manager"], course["category"]]
        total = 0
        for account in account_order:
            executed = by_course_account.get((course["id"], account), [0, 0, 0])[2]
            row.append(executed)
            total += executed
        row.append(total)
        course_account_matrix.append(row)

    raw_export = [
        [
            row["source_row"],
            row["date"],
            row["course"],
            row["account"],
            row["item"],
            row["amount"],
            row["vendor"],
            row["resolution_no"],
            row["confidence"],
            row["content"],
        ]
        for row in raw_rows
    ]
    unmatched = [row for row in raw_export if row[8] == "확인필요"]

    xlsx_path = OUTPUT_DIR / "지출내역_예산프로그램_정리본_20260702.xlsx"
    out = Workbook()
    write_sheet(out.active, course_rows, ["과정", "담당", "구분", "원예산", "조정예산", "집행액", "잔액", "집행률(%)"])
    out.active.title = "과정별_요약"
    write_sheet(out.create_sheet("세목별_요약"), account_rows, ["세목", "원예산", "조정예산", "집행액", "잔액", "집행률(%)"])
    write_sheet(out.create_sheet("과정별_세목별"), course_account_rows, ["과정", "담당", "구분", "세목", "원예산", "조정예산", "집행액", "잔액", "집행률(%)"])
    write_sheet(out.create_sheet("과정세목_매트릭스"), course_account_matrix, ["과정", "담당", "구분", *account_order, "합계"])
    write_sheet(out.create_sheet("항목별_반영"), summary_rows, ["과정", "담당", "구분", "세목", "항목", "원예산", "조정예산", "집행액", "잔액", "집행률(%)"])
    write_sheet(out.create_sheet("원장_매칭"), raw_export, ["원본행", "지출일", "과정", "세목", "항목", "금액", "사용처", "결의번호", "신뢰도", "내용"])
    write_sheet(out.create_sheet("미분류_확인필요"), unmatched, ["원본행", "지출일", "과정", "세목", "항목", "금액", "사용처", "결의번호", "신뢰도", "내용"])
    out.save(xlsx_path)

    matched_total = sum(row["amount"] for row in raw_rows if row["item_id"])
    unmatched_total = sum(row["amount"] for row in raw_rows if not row["item_id"])
    print(json.dumps({
        "expenseFile": str(expense_path),
        "json": str(json_path),
        "xlsx": str(xlsx_path),
        "rows": len(raw_rows),
        "matchedRows": len(executions),
        "total": sum(row["amount"] for row in raw_rows),
        "matchedTotal": matched_total,
        "unmatchedTotal": unmatched_total,
        "unmatchedRows": len(raw_rows) - len(executions),
        "courseTop": sorted(course_rows, key=lambda row: row[5], reverse=True)[:8],
        "accountRows": account_rows,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
